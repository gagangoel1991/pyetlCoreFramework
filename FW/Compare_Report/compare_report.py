import sys, os, shutil, openpyxl, threading, multiprocessing,  time, datetime, pytz, getpass
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook 
from openpyxl.styles import PatternFill
import FW.Initialize.initialize_global_variables as iniVar
from openpyxl.styles import Border, Side, Alignment, Font, Color
from FW.FW_logger import loggerPass, loggerFail, loggerInfo, loggerDisplay
from FW.FW_logger import add_in_reporting_dict, check_key_in_reporting_dict, get_from_reporting_dict, add_detail_tabs_info_in_reporting_dict
import FW.FW_logger as logger
from FW.FW_exec_db_update import update_exec_db
from FW.FW_Run_Post_Hook import run_post_test_hook

# Max diffrence records that will be populated in difference dataframes by default 
default_reportdiffSize = 1000
def set_data_size_in_report(size):
    """Override the number of records to be printed in Excel report for differring and extra records"""
    if str(size).lower() == 'all':
        add_in_reporting_dict('data_size_in_report', 100000000)
    else:
        size = size if size % 2 == 0 else size+1
        add_in_reporting_dict('data_size_in_report', size)

def set_custom_name_for_result_report(repName):
    """This sets the custom name of result report. Can be called from main function in script"""
    add_in_reporting_dict('custom_name_for_result_report', repName)

def get_custom_name_set_for_result_report():
    """This gets the custom name of result report. """
    return get_from_reporting_dict('custom_name_for_result_report') if \
                                check_key_in_reporting_dict('custom_name_for_result_report') else "report"

def getAndSaveDuplicateReferenceValuesInDF(df,ls_ref_cols, save_csv_name='duplicate_ref_values'):
    """this will dump the duplicate values in ref columns"""
    df = df.reset_index()[ls_ref_cols].sort_values(ls_ref_cols)
    df_dup= df[df.duplicated(ls_ref_cols, keep=False)].groupby(ls_ref_cols).size().reset_index().rename(columns={0:'Count'})

    save_csv_path = os.path.join(iniVar.current_project_path, "Reports", save_csv_name + ".csv")
    df_dup.to_csv(save_csv_path, index=False)
    loggerPass(f"Duplicate values in ref column written to '{save_csv_path}' successfully")

    return df_dup

def compare(ls_src_info, ls_trg_info, report_tab_name =None, numeric_threshold = 0):
    """Compares two dataframes and generates diffence dataframe, extra rows in first dataframe, extra rows in second dataframe. Additionally this function will add many runtime information to the reporting dictionary that will be used for report preparation.

    This functions compares given 2 dataframes and stores the results into reporting
    dictionary along with some attributes of result dataframes.

    Parameters
    ----------
    ls_src_info : list
        List containing following information about first dataframe
        1. First DataFrame 2. List of Reference columns 3. List of columns to compare
    ls_trg_info : list
        List containing following information about second dataframe
        1. Second DataFrame 2. List of Reference columns 3. List of columns to compare
    report_tab_name : str, default None
        If not null, then store the comparison table result in this tab in result report
    numeric_threshold : int or float, default 0
        While comparing 2 data sets, for numerical data, it will not fail if the difference is less than or equal to threshold value specified.

    returns
    --------
    None

    examples
    --------
    compare([df_src_tr,['POLICY_NUMBER_TEXT','COVERAGE_NUMBER_TEXT'],['POLICY_NUMBER_TEXT','COVERAGE_NUMBER_TEXT','GENDER_CODE']],
                            [df_trg_tr,['POLICY_NUMBER_TEXT','COVERAGE_NUMBER_TEXT'],['POLICY_NUMBER_TEXT','COVERAGE_NUMBER_TEXT','GENDER_CODE']], report_tab_name = "Tab12")
    """

    reportdiffSize = get_from_reporting_dict('data_size_in_report') if check_key_in_reporting_dict('data_size_in_report') else default_reportdiffSize

    #[df_src_tr,['POLICY_NUMBER_TEXT','COVERAGE_NUMBER_TEXT'],['POLICY_NUMBER_TEXT','COVERAGE_NUMBER_TEXT','GENDER_CODE']],
    #[df_trg_tr,['POLICY_NUMBER_TEXT','COVERAGE_NUMBER_TEXT'],['POLICY_NUMBER_TEXT','COVERAGE_NUMBER_TEXT','GENDER_CODE']])
    report_tab_name = report_tab_name.replace("-", "_").replace(" ", "") if report_tab_name != None else None

    # Parse info from list arg
    src_df = ls_src_info[0]
    src_df.columns = [x.upper() for x in src_df.columns]
    ls_ref_src = [x.upper() for x in ls_src_info[1]]
    ls_col_to_comp_src = [x.upper() for x in ls_src_info[2]]

    trg_df = ls_trg_info[0]
    trg_df.columns = [x.upper() for x in trg_df.columns]
    ls_ref_trg = [x.upper() for x in ls_trg_info[1]]
    ls_col_to_comp_trg = [x.upper() for x in ls_trg_info[2]]

    # required columns only and make type to str for comparison
    # src_df = src_df[ls_col_to_comp_src].copy()
    # trg_df = trg_df[ls_col_to_comp_trg].copy()

    # src_df = src_df[ls_col_to_comp_src].copy().astype(str)
    # trg_df = trg_df[ls_col_to_comp_trg].copy().astype(str)

    src_df = src_df[ls_col_to_comp_src].copy().astype(object).astype(str)
    trg_df = trg_df[ls_col_to_comp_trg].copy().astype(object).astype(str)

    if report_tab_name == None:
        _update_dict_for_summary_record_from_compare_function(src_df, trg_df, ls_ref_src)

    #### all column names in 2 lists should match else error to be raised.==================
    if set(ls_col_to_comp_src) == set(ls_col_to_comp_trg):
        loggerInfo(f"columns to be validated are {ls_col_to_comp_trg} and are same in source and target")
    else:
        raise Exception(f"columns to be validated in source and target are different. In source, {ls_col_to_comp_src} and in target {ls_col_to_comp_trg}")

    # Change index to numeric tpye if possible
    for c in ls_ref_src:
        if (src_df[c].astype(str).str.isnumeric().all() == True) and (trg_df[c].astype(str).str.isnumeric().all() == True):
            src_df[c] = src_df[c].astype('int64')
            trg_df[c] = trg_df[c].astype('int64')

    src_df = src_df.set_index(ls_ref_src).sort_index()
    trg_df = trg_df.set_index(ls_ref_trg).sort_index()

    #### check uniquie ness of reference columns
    _reportUniquenessOfReferenceCols(src_df,trg_df )

    # Comparison threshold info
    add_in_reporting_dict('numeric_threshold', numeric_threshold)
    if numeric_threshold > 0:
        loggerInfo(f"Numerical values comparison is done with a difference threshold of <= {numeric_threshold} as acceptable")

    # common diffs
    comm_diffs = _getCommonDiffs(src_df, trg_df, ls_ref_src, numeric_threshold )  #Assuming col names in src and trg are same
    comm_diffs = _get_diffs_with_diffs_on_top(comm_diffs)  # bring diff records in all columns on top
    # extra in src
    extra_src  = src_df[~src_df.index.isin(trg_df.index)].reset_index()
    # extra in trg
    extra_trg = trg_df[~trg_df.index.isin(src_df.index)].reset_index()

    # dfs to be used in reporting
    if report_tab_name == None:
        _update_dict_with_diffs_records_for_summary_from_compare_function(comm_diffs, extra_src, extra_trg)

        add_in_reporting_dict('comm_diffs', comm_diffs[:reportdiffSize])
        add_in_reporting_dict('extra_src', extra_src[:reportdiffSize])
        add_in_reporting_dict('extra_trg', extra_trg[:reportdiffSize])

        if int(get_from_reporting_dict('mismatch_cnt')) == 0:
            loggerPass(f"Total mismatch : {int(get_from_reporting_dict('mismatch_cnt'))}, columns having differences along with counts : {get_from_reporting_dict('mismatch_cols_n_cnt')}")
        else:
            loggerFail(
                f"Total mismatch : {int(get_from_reporting_dict('mismatch_cnt'))}, columns having differences along with counts : {get_from_reporting_dict('mismatch_cols_n_cnt')}")

        if get_from_reporting_dict('extra_src_cnt')==0 and get_from_reporting_dict('extra_trg_cnt')==0:
            loggerPass(f"Extra rows in src : {get_from_reporting_dict('extra_src_cnt')}, and extra rows in trg : {get_from_reporting_dict('extra_trg_cnt')}")
        else:
            loggerFail(
                f"Extra rows in src : {get_from_reporting_dict('extra_src_cnt')}, and extra rows in trg : {get_from_reporting_dict('extra_trg_cnt')}")

    #'detail_tabs' will be created
    if report_tab_name != None:
        if int(len(comm_diffs)/2)==0:
            loggerPass( f"Total mismatch : {int(len(comm_diffs)/2)}, columns having differences along with counts : {_getTotalDiffColumns(comm_diffs)}")
        else:
            loggerFail(
                f"Total mismatch : {int(len(comm_diffs) / 2)}, columns having differences along with counts : {_getTotalDiffColumns(comm_diffs)}")

        if len(extra_src)==0 and len(extra_trg)==0:
            loggerPass(f"Extra rows in src : {len(extra_src)}, and extra rows in trg : {len(extra_trg)}")
        else:
            loggerFail(f"Extra rows in src : {len(extra_src)}, and extra rows in trg : {len(extra_trg)}")

        add_detail_tabs_info_in_reporting_dict(report_tab_name, [comm_diffs[:reportdiffSize], extra_src[:reportdiffSize], extra_trg[:reportdiffSize]])
        if len(comm_diffs) ==0 and len(extra_src) ==0 and len(extra_trg) ==0:
            loggerPass(f"Comparison done successfully and ONLY the differences found are written in report tab {str(threading.get_ident()) + '-' + report_tab_name} ")
        else:
            loggerFail(f"Comparison mismatches found, only the differences are written in report tab {str(threading.get_ident()) + '-' + report_tab_name}")

def prepareReport(rptCnt, testName, Error, totaltestsCount, reportName="report", replace_spaces_with_star=True):
    """Prepares report of single or multi-script execution. This is part of test script template. It takes various information to
    be included in report from reporting dictionary internally.

    Parameters
    ----------
    rptCnt : int
        Internally used to check if all the executions completed in multi-script execution
    testname : string
        Name of current test script
    Error : string
        Error message if some error occoured in test_main() function of any script
    totaltestsCount : int
        In case of multi-script execution, this is count of total automation scripts in batch
    reportName : string, defautl 'report.xlsx'
        Name of the report to be generated in Report folder of project
    replace_spaces_with_star : bool, default Ture
        if True, it replaces all spaces with * in result report
    returns
    --------
    None

    """
    reportdiffSize = get_from_reporting_dict('data_size_in_report') if check_key_in_reporting_dict(
        'data_size_in_report') else default_reportdiffSize

    df_cd =  get_from_reporting_dict('comm_diffs') if check_key_in_reporting_dict('comm_diffs') else None
    df_x1 =  get_from_reporting_dict('extra_src') if check_key_in_reporting_dict('extra_src') else None
    df_x2 =  get_from_reporting_dict('extra_trg') if check_key_in_reporting_dict('extra_trg') else None
    
    report_template = os.path.join(iniVar.current_project_path, "Resources", "report_template.xlsx")

    if totaltestsCount==1:
        report_file_path = os.path.join(iniVar.current_project_path, "Reports", reportName + ".xlsx")
    else:
        report_file_path = os.path.join(iniVar.current_project_path, "Reports", "report.xlsx")
            
    # adding report file path to reporting dictionary
    add_in_reporting_dict('report_path', report_file_path)
    # 1 . copy report to report location with current name only once
    if rptCnt == 1: shutil.copy(report_template, report_file_path)
    # 2. open report
    wb = load_workbook(report_file_path)
    writer = pd.ExcelWriter(report_file_path, engine='openpyxl')
    writer.book = wb
    s_sh = wb["Summary"]
    
    # This will addd new sheet
    nameofNewSheet = "Script_logs_" + str(_get_first_blank_row(s_sh, col = 'A') - 6)
    t_sh = wb.create_sheet(nameofNewSheet)
    writer.sheets = dict((ws.title, ws) for ws in wb.worksheets)
    
    # Write summary record
    s_sh = wb["Summary"]
    _update_dict_for_summary_record_from_report_function(df_cd, df_x1, df_x2, testName, nameofNewSheet, Error )
    _enter_info_in_summary_result(s_sh, writer)
    
    
    # =====================write detailed report tab======================
    if Error == None and repr(type(df_cd)) != "<class 'NoneType'>" and repr(type(df_x1)) != "<class 'NoneType'>" and repr(type(df_x2)) != "<class 'NoneType'>":
    # write common diffs Replace all blanks with * and export
        if replace_spaces_with_star == True:
            df_cd.replace(r'\s','*',regex=True,inplace=True)
        df_cd = _write_in_detailed_report(t_sh, 1, 'Common Diffs', df_cd, writer, nameofNewSheet, 1, 2,
                                          f'Showing only first {int(reportdiffSize/2)} mismatches…' )  # msg wont be written
        writer.save()
        # write Extra in source
        rowtowrite = len(df_cd) + 4
        df_x1 = _write_in_detailed_report(t_sh, rowtowrite, 'Extra in Source', df_x1, writer, nameofNewSheet,
                                          rowtowrite, 3, f'Showing only first {reportdiffSize} records…' )  # msg wont be written
        writer.save()
        # write Extra in target
        rowtowrite = len(df_cd) + len(df_x1) + 7
        df_x2 = _write_in_detailed_report(t_sh, rowtowrite, 'Extra in Target', df_x2, writer, nameofNewSheet,
                                          rowtowrite, 3, f'Showing only first {reportdiffSize} records…' )
        writer.save()
        #fetch coordinates of diff cells - used later
        
        df_cd = df_cd.set_index('index')
        ls_diffCells = _getDiffCellsCoordinates(df_cd)
        #format the diff tab
        _formatExcelCellColor(t_sh, ls_diffCells)
        writer.save()

    #Add additonal detail tabs info and add that in report
    if logger.check_detail_tabs_info_present_in_reporting_dict() == True:
        _add_detail_tabs_info_in_report(writer, wb)

    # write logger steps in detail tab
    _write_Steps_Info_Log_in_Detailed_Sheet(t_sh)
    writer.save()

    # store variable in dict for updating metrics
    add_in_reporting_dict('test_progress_status', 'End')
    add_in_reporting_dict('end_time', str(datetime.now(pytz.timezone('America/Toronto')))[:23])

    # execution time taken
    DTStarTime = get_from_reporting_dict('start_time')
    DTEndTime = get_from_reporting_dict('end_time')
    RunTime = (datetime.strptime(DTEndTime, '%Y-%m-%d %H:%M:%S.%f') - datetime.strptime(DTStarTime,'%Y-%m-%d %H:%M:%S.%f')).seconds
    add_in_reporting_dict('running_time', RunTime)

    # write logger steps in detail tab for running time update
    _write_Running_Time_Info_Log_in_Detailed_Sheet(t_sh)
    writer.save()
    writer.close()

    # Updade metrics
    if iniVar.dry_run == False:
        update_exec_db()
        #pass
    else:
        print("Dry-run = True, no auto-execution-db updated")

    # run_post_test_hook will run for only those functions that are required to be executed in
    # post hook and where post_test_hook_function_list is not 'none'.
    post_test_hook_function_list = get_from_reporting_dict('post_test_hook_function_list') if check_key_in_reporting_dict('post_test_hook_function_list') else 'None'
    if post_test_hook_function_list != 'None':
        post_test_hook_function_parameter_list = get_from_reporting_dict('post_test_hook_function_parameter_list') if check_key_in_reporting_dict('post_test_hook_function_parameter_list') else None
        run_post_test_hook(post_test_hook_function_list, post_test_hook_function_parameter_list)

    # Every script execution adds meta data in global dict
    iniVar._set_test_info_in_global_dict(rptCnt)

    # open report
    if totaltestsCount ==rptCnt:
        loggerInfo("Report generated successfully, now opening it")
        os.startfile(report_file_path)
        loggerInfo("Report opened successfully")

        # if iniVar.dry_run == False:  # will be changed to false
        #     import FW.QTest_Integration.FW_pulse_integration as pulseInt
        #      pulseInt.send_result_data_to_qtest(report_file_path)

            
            
def _get_diffs_with_diffs_on_top(df, num_of_diff_record_on_top =5):
    '''Rearrange the diffs dataframe such that we have atmost some failed records from each column in top records in report'''

    from collections import defaultdict
    coord = _getDiffCellsCoordinates(df)                # cell coordinates where mismatch occured are fetched.
    d = defaultdict(list)
    for v, k in coord:
        d[k].append(v) if k != 2 else None  #avoid column index

    ls = []
    _ = [ls.extend(a) for a in [x[:num_of_diff_record_on_top] for x in d.values()]]
    ls.extend([x - 1 for x in ls])
    rowlist = sorted(list(set(ls)))

    df_top = df.iloc[rowlist]
    df_top = df_top.append(df[~df.index.isin(df_top.index)].dropna())              # rest of the records from df will be added to df_top after df_top has max no of difference records from each column on top.
    return df_top

def _changeDataToCatagory(df, bool_reduce_size = True):
    """Reduces the size of dataframe"""
    if len(df) > 50000 and bool_reduce_size==True:
        print(f"Optimizing data set size...")
        a = time.time()
        ini_size = df.memory_usage(index=True, deep=True).sum() / (1024*1024*1024)
        objcols = df.select_dtypes(['object']).columns

        for col in objcols:
            # if column is pure date column, avoid that from optimization
            try:
                var_isDate = False
                if ((pd.to_datetime(df[col]).dt.floor('d')==pd.to_datetime(df[col])).all()):
                    var_isDate = True
            except:
                pass

            if var_isDate==False:
                if df[col].astype('category').memory_usage() < df[col].memory_usage():
                    df[col] = df[col].astype('category', copy=False)

        b = time.time()
        final_size = df.memory_usage(index=True, deep=True).sum() / (1024 * 1024 * 1024)
        print(f"Data set size optimized from {round(ini_size,4)} GB to {round(final_size,4)} GB in {round((b-a)/60,4)} minutes")
    return df


def _add_detail_tabs_info_in_report(writer, wb):
    """Add custom additional tab having details in result file"""

    reportdiffSize = get_from_reporting_dict('data_size_in_report') if check_key_in_reporting_dict(
        'data_size_in_report') else default_reportdiffSize

    lst_tab_names = logger.get_detail_tabs_name_list_from_reporting_dict()
    add_in_reporting_dict('list_tab_name', lst_tab_names)

    # Add renamed tab names for additional detaild info in reporting dict
    dSName = logger.get_from_reporting_dict('detailSheetName')
    #replacing the thread id of detailed sheet name with the first and last character of dSName.
    lst_renamed_tab_names = [name.replace(str(threading.get_ident()),dSName[0]+dSName[-1]) for name in lst_tab_names]
    logger.add_in_reporting_dict("lst_renamed_tab_names", lst_renamed_tab_names)

    for nameofNewSheet, nameOfRenamedSheet in zip(lst_tab_names, lst_renamed_tab_names):
        tab_name_info_lst = logger.get_detail_tabs_info_values_list_from_reporting_dict(nameofNewSheet)

        #Add custom renamed details sheet tab
        t_sh = wb.create_sheet(nameOfRenamedSheet)  # Add renamed name
        writer.sheets = dict((ws.title, ws) for ws in wb.worksheets)

        # Write common diffs
        df_cd = tab_name_info_lst[0]
        df_cd.replace(r'\s', '*', regex=True, inplace=True)
        df_cd = _write_in_detailed_report(t_sh, 1, 'Common Diffs', df_cd, writer, nameOfRenamedSheet, 1, 2,
                                          f'Showing only first {int(reportdiffSize/2)} mismatches…')   # msg wont be written
        writer.save()

        # write Extra in source
        rowtowrite = len(df_cd) + 4
        df_x1 = tab_name_info_lst[1]
        df_x1 = _write_in_detailed_report(t_sh, rowtowrite, 'Extra in Source', df_x1, writer, nameOfRenamedSheet, rowtowrite, 3,
                                          f'Showing only first {reportdiffSize} records…')  # msg wont be written
        writer.save()

        # write Extra in target
        rowtowrite = len(df_cd) + len(df_x1) + 7
        df_x2 = tab_name_info_lst[2]
        df_x2 = _write_in_detailed_report(t_sh, rowtowrite, 'Extra in Target', df_x2, writer, nameOfRenamedSheet, rowtowrite, 3,
                                          f'Showing only first {reportdiffSize} records…')   # msg wont be written
        writer.save()

        # fetch coordinates of diff cells - used later
        df_cd = df_cd.set_index('index')
        ls_diffCells = _getDiffCellsCoordinates(df_cd)
        # format the diff tab
        _formatExcelCellColor(t_sh, ls_diffCells)
        writer.save()


def _write_Steps_Info_Log_in_Detailed_Sheet(t_sh):
    """Writes run time loggers in detailed sheet of report"""

    stepList = iniVar.th_local.dict['logger']
    thId = str(threading.get_ident())
    stepList = [x[(len(thId)+3) :] for x in stepList if thId in x]
    
    col = _get_first_blank_column(t_sh, row = 1) + 2
    if len(stepList)>0:  t_sh.cell(1, col).value = "Logger Steps:"
    for i, s in enumerate(stepList):
        #writing logger info in sheet.
        t_sh.cell(i+3, col).value = stepList[i]
        _format_logger_steps_in_detailed_tab(t_sh.cell(i+3, col) )

    st = get_from_reporting_dict('overall_status').upper()
    t_sh.cell(i + 5, col).value = f"Overall test status = {st}"


def _write_Running_Time_Info_Log_in_Detailed_Sheet(t_sh):
    """Writes total running time log in detailed sheet of report"""
    scol = _get_first_blank_column(t_sh, row=1) -1
    scol_letter = openpyxl.utils.get_column_letter(scol)
    srow = _get_last_non_blank_row(t_sh, col=scol_letter) + 1
    t_sh.cell(srow, scol).value = f"Script execution time = {get_from_reporting_dict('running_time')} sec"


def _format_logger_steps_in_detailed_tab(cell):
    """format steps having reference to additional detailed tab in blue link """

    lst_renamed_tab_names = get_from_reporting_dict('lst_renamed_tab_names') if check_key_in_reporting_dict('lst_renamed_tab_names') else None
    ls_tab_name = get_from_reporting_dict('list_tab_name') if check_key_in_reporting_dict('list_tab_name') else None

    # Format colors of pass and fail normal steps
    if cell.value[:4]=='PASS':
        cell.font = Font(color='FF008000', bold=True)
    if cell.value[:4]=='FAIL':
        cell.font = Font(color='FFFF0000', bold=True)

    # Format steps having additional information
    # if ls_tab_name != None and len([tab for tab in ls_tab_name[::-1] if tab in cell.value]) > 0:
    #     tab_name = [tab for tab in ls_tab_name[::-1] if tab in cell.value][0]
    # Below code implemented to fix the bug which causes wrong hyperlinks for tabs having couunt >10

    if ls_tab_name != None and len(_check_additional_tab_name_in_logger(cell.value)) > 0:
        tab_name = _check_additional_tab_name_in_logger(cell.value)
        # Replace the tab name in the logger step with renamed tab name
        renamed_tab_name = lst_renamed_tab_names[ls_tab_name.index(tab_name)]
        renamed_logger_step = cell.value.replace(tab_name,renamed_tab_name )

        # Write renamed logger step in detailed tab
        cell.hyperlink = "#'" + renamed_tab_name + "'!A1"
        cell.value = renamed_logger_step

        if cell.value[:4] == 'PASS':
            cell.font = Font(italic=True, underline='single', color='FF008000')
        if cell.value[:4] == 'FAIL':
            cell.font = Font(italic=True, underline='single', color='FFFF0000')

def _check_additional_tab_name_in_logger(cell_value):
    """returns the additional detail tab name if present in current cell logger"""
    try:
        ret = cell_value[cell_value.index(str(threading.get_ident()) + '-'):].strip()
    except:
        ret=''
    return ret


def _update_dict_for_summary_record_from_compare_function(src_df, trg_df, ls_ref_src):
    '''Add original datasets info in reporting dictionary'''

    ls_cols_validate = [x for x in (set(src_df.columns) - set(ls_ref_src))]
    col_names =',\n'.join(ls_cols_validate)
    add_in_reporting_dict('col_names', col_names)

    size_of_src = str(len(src_df)) + ' x ' + str(len(src_df.columns) - len(ls_ref_src))
    add_in_reporting_dict('size_of_src', size_of_src)

    size_of_trg = str(len(trg_df)) + ' x ' + str(len(trg_df.columns) - len(ls_ref_src))
    add_in_reporting_dict('size_of_trg', size_of_trg)

    unique_cnt = ';\n'.join([c + ': src=' + str(src_df[c].nunique()) +  ' trg=' + str(trg_df[c].nunique()) for c  in ls_cols_validate])
    add_in_reporting_dict('unique_cnt', unique_cnt)

    null_cnt = ';\n'.join([c + ': src=' + str(len(src_df[src_df[c]=='(null)'])) +  ' trg=' + str(len(trg_df[trg_df[c]=='(null)'])) for c in ls_cols_validate])
    add_in_reporting_dict('null_cnt', null_cnt)

def _update_dict_with_diffs_records_for_summary_from_compare_function(df_cd, df_x1, df_x2):
    """Dictionary is updated with information of differences like mismatch in col and their count, extra records in source and target if any"""

    mismatch_cnt = len(df_cd.index) / 2 if repr(type(df_cd)) != "<class 'NoneType'>" else None
    add_in_reporting_dict('mismatch_cnt', mismatch_cnt)
    extra_src_cnt = len(df_x1.index) if repr(type(df_x1)) != "<class 'NoneType'>" else None
    add_in_reporting_dict('extra_src_cnt', extra_src_cnt)
    extra_trg_cnt = len(df_x2.index) if repr(type(df_x2)) != "<class 'NoneType'>" else None
    add_in_reporting_dict('extra_trg_cnt', extra_trg_cnt)
    mismatch_cols_n_cnt = _getTotalDiffColumns(df_cd) if repr(type(df_cd)) != "<class 'NoneType'>" else None
    add_in_reporting_dict('mismatch_cols_n_cnt', mismatch_cols_n_cnt)

def _update_dict_for_summary_record_from_report_function(df_cd, df_x1, df_x2, testName, detailSheetName, Error = None):
    """Add differences info or other info available during report function, into reporting dictionary for later use"""

    #if Error == None:
        # mismatch_cnt= len(df_cd.index)/2 if repr(type(df_cd)) != "<class 'NoneType'>" else None
        # add_in_reporting_dict('mismatch_cnt', mismatch_cnt)
        # extra_src_cnt= len(df_x1.index) if repr(type(df_x1)) != "<class 'NoneType'>" else None
        # add_in_reporting_dict('extra_src_cnt', extra_src_cnt)
        # extra_trg_cnt= len(df_x2.index) if repr(type(df_x2)) != "<class 'NoneType'>" else None
        # add_in_reporting_dict('extra_trg_cnt', extra_trg_cnt)
        # mismatch_cols_n_cnt = _getTotalDiffColumns(df_cd) if repr(type(df_cd)) != "<class 'NoneType'>" else None
        # add_in_reporting_dict('mismatch_cols_n_cnt', mismatch_cols_n_cnt)
    
    add_in_reporting_dict('testName', testName)
    add_in_reporting_dict('detailSheetName', detailSheetName)
    add_in_reporting_dict('error', Error)


def _enter_info_in_summary_result(s_sh, writer):
    """Enter information in summary sheet row based on info available in reporting dictionary"""

    status = "Pass"
    vRow = str(_get_first_blank_row(s_sh))
    s_sh['A' + vRow] = int(vRow) - 6
    s_sh['B' + vRow] = get_from_reporting_dict('testName')

    if check_key_in_reporting_dict('error') and (get_from_reporting_dict('error')!= None):  # when error
        s_sh['D' + vRow] = get_from_reporting_dict('error') 
        status = "Fail"
        wrapText = False
    else:
        wrapText = True
        s_sh['D' + vRow] = get_from_reporting_dict('col_names') if check_key_in_reporting_dict('col_names') else None
        s_sh['E' + vRow] = get_from_reporting_dict('size_of_src') if check_key_in_reporting_dict('size_of_src') else None
        s_sh['F' + vRow] = get_from_reporting_dict('size_of_trg') if check_key_in_reporting_dict('size_of_trg') else None
        s_sh['G' + vRow] = get_from_reporting_dict('mismatch_cnt') if check_key_in_reporting_dict('mismatch_cnt') else None
        s_sh['H' + vRow] = get_from_reporting_dict('mismatch_cols_n_cnt') if check_key_in_reporting_dict('mismatch_cols_n_cnt') else None
        s_sh['I' + vRow] = get_from_reporting_dict('extra_src_cnt') if check_key_in_reporting_dict('extra_src_cnt') else None
        s_sh['J' + vRow] = get_from_reporting_dict('extra_trg_cnt') if check_key_in_reporting_dict('extra_trg_cnt') else None
        s_sh['K' + vRow] = get_from_reporting_dict('unique_cnt') if check_key_in_reporting_dict('unique_cnt') else None
        s_sh['L' + vRow] = get_from_reporting_dict('null_cnt') if check_key_in_reporting_dict('null_cnt') else None

    writer.save()
            
    if _getOverallStatus(s_sh, vRow) == "Fail":
        status = "Fail"

    add_in_reporting_dict('overall_status', status)

    if status == "Fail":
        print('')
        loggerFail("Overall test script status is FAILED")
        s_sh['B' + vRow].fill = PatternFill(start_color='FF0000', end_color='FF0000',fill_type='solid')
        # in case more records to write to report, then below will show warning msg
        _print_wait_info_for_preparing_Report()
    else:
        s_sh['B' + vRow].fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type = "solid")
        print('')
        loggerPass("Overall test script status is PASSED")

    if check_key_in_reporting_dict('detailSheetName'):
        shname = get_from_reporting_dict("detailSheetName")
        cell = s_sh['C' + vRow]
        cell.hyperlink = "#" + shname + "!A1"   
        cell.value = "Detailed Report Link"
        cell.font = Font(italic=True, underline = 'single', color = 'FF0000FF')

    # format the cells in the second row
    for cell in s_sh[vRow + ":" + vRow]:
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True,)
        thin = Side(border_style="thin", color="000000")
        cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
        
    # if error, dont wrap error msg
    if wrapText == False:
        s_sh['D' + vRow].alignment = Alignment(horizontal="left", vertical="center", wrap_text=False,)

def _print_wait_info_for_preparing_Report():
    ifPrimaryTabDFPresent = True if check_key_in_reporting_dict('comm_diffs') else False
    ifAdditionalTabDFPresent = logger.check_detail_tabs_info_present_in_reporting_dict()

    if (ifPrimaryTabDFPresent==False) and (ifAdditionalTabDFPresent==False):
        return

    reportdiffSize = get_from_reporting_dict('data_size_in_report') if check_key_in_reporting_dict(
        'data_size_in_report') else default_reportdiffSize

    DTStarTime = get_from_reporting_dict('start_time')
    CurrTime = str(datetime.now(pytz.timezone('America/Toronto')))[:23]
    elapsedTime = (datetime.strptime(CurrTime, '%Y-%m-%d %H:%M:%S.%f') - datetime.strptime(DTStarTime,
                                                                                           '%Y-%m-%d %H:%M:%S.%f')).seconds

    diffSize = 'ALL' if reportdiffSize == 100000000 else reportdiffSize

    if diffSize == 'ALL':
        loggerDisplay()
        loggerDisplay(f"WARNING - Script completed except report generation in time : {elapsedTime} sec")
        loggerDisplay(
            f"However, as the user selected to see 'ALL' the extra and differring records in Report,")
        loggerDisplay("It may take FEW MINUTES to generate report depending on the size of data to write in report")
        loggerDisplay(
            "It is recommended not to use 'set_data_size_in_report(size)' with size > 1000, due to performance issues in excel report writing, unless very much required.")
        print("Please wait for report generation...")
        loggerDisplay()

    elif diffSize > default_reportdiffSize:
        loggerDisplay()
        loggerDisplay(f"WARNING - Script completed except report generation in time : {elapsedTime} sec")
        loggerDisplay(f"However, as the user selected to see maximum of {reportdiffSize} records for the extra and differring records in Report,")
        loggerDisplay(f"it may take FEW SECONDS to FEW MINUTES to generate report depending of size selected.")
        loggerDisplay(
            "It is recommended not to use 'set_data_size_in_report(size)' with size > 1000, due to performance issues in excel report writing, unless very much required.")
        print("Please wait for report generation...")
        loggerDisplay()

def _getOverallStatus(s_sh, vRow):
    """Determine the status of script based on logger steps"""

    status = "Pass"
    stepList = iniVar.th_local.dict['logger']
    thId = str(threading.get_ident())
    stepList = [x[(len(thId)+3) :] for x in stepList if thId in x]
    jointStepList = ",".join(stepList)
    
    if "FAIL -" in jointStepList:
        status = "Fail"
        
    #check summary values for status
    if ((s_sh['G' + vRow].value != None and s_sh['G' + vRow].value!=0) or 
        (s_sh['I' + vRow].value != None and s_sh['I' + vRow].value!=0) or 
        (s_sh['J' + vRow].value != None and s_sh['J' + vRow].value!=0)):
        
            status = "Fail"
        
    # check for uniques


    if s_sh['K' + vRow].value != None:
        import re
        unqCntStr = s_sh['K' + vRow].value
        lst = re.findall("=\d+", unqCntStr)
        for i in range(1, len(lst),2):
            if lst[i]!=lst[i-1] :
                # if threshold > 0, then number of uniques can be different and has to be ignored
                numeric_threshold = get_from_reporting_dict('numeric_threshold') if check_key_in_reporting_dict('numeric_threshold') else 0
                if numeric_threshold == 0:
                    status = "Fail"

    # check for nulls
    if s_sh['L' + vRow].value != None:
        import re
        unqCntStr = s_sh['L' + vRow].value
        lst = re.findall("=\d+", unqCntStr)
        for i in range(1, len(lst),2):
            if lst[i]!=lst[i-1] :
                status = "Fail"

    return status


def _get_first_blank_row(s_sh, col = 'A'):
    """Find first blank row in sheet"""

    i = 1
    lst = len(s_sh[col])
    for i in range(lst,1,-1):
        if s_sh[col + str(i)].value != '':
            return i+1

def _get_last_non_blank_row(s_sh, col = 'A'):
    """Find last no blank row in sheet in given column"""

    lst = len(s_sh[col])+1
    for i in range(lst,1,-1):
        if s_sh[col + str(i)].value != None:
            return i+1

def _get_first_blank_column(sh, row = 1):
    """Find first blank column in the sheet"""

    i = 1
    lst = len(sh[row])
    if lst == 1 : return 1
    
    for i in range(lst,1,-1):
        if sh.cell(row, i).value != '':
            return i+1


def _getTotalDiffColumns(df_cd_full):
    """Gets the name of columns and counts of differences in each column"""
    if len(df_cd_full)!=0:
        numeric_threshold = get_from_reporting_dict('numeric_threshold') if check_key_in_reporting_dict('numeric_threshold') else 0
        if numeric_threshold == 0:
            shiftedDiffs = df_cd_full.shift() == df_cd_full
            diffCounts= shiftedDiffs[shiftedDiffs.reset_index().index % 2 != 0].apply(lambda x: x.value_counts(),axis = 0)
        else: # if thresold is there
            diffCounts = apply_numerical_threshold(df_cd_full, numeric_threshold).apply(lambda x: x.value_counts(),axis = 0)

        ggg = diffCounts.reset_index(drop=True).loc[0].dropna().to_frame().sort_values(0,ascending=False)
        ls_Tuples =  [(x,int(y)) for x,y in zip(ggg.index.tolist(),ggg[0])]
        shiftedDiffs,diffCounts = [None]*2
        return str([(x,y) for x,y in ls_Tuples if x != 'index'])[1:-1]
    else:
        return 'None'


def isnumber(x):
    # checks if argument is number or not
    try:
        float(x)
        return True
    except ValueError:
        return False

def diff_for_2_vals(vals, threshold):
    # for numerical values, checks if difference of two values is greater or equal to threshold or not
    x = vals.iloc[0]
    y = vals.iloc[1]
    if isnumber(x) and isnumber(y):
        if abs(float(x) - float(y)) <= threshold:
            return True
        else:
            return False
    else:
        return True if x == y else False


def compare_apply_numerical_threshold(df1, threshold):
    #
    if threshold==0:
        return df1

    #df1['seq_custom_intenal'] = [x for x in range(0, len(df1))]   # adding
    # df1 = df1.reset_index(drop=True)
    #
    # df11 = df1[df1.index % 2 != 0].drop(columns=['seq_custom_intenal', 'index'])
    #
    # df22 = df1.shift()
    # df22 = df22[df22.index % 2 != 0].drop(columns=['seq_custom_intenal', 'index'])
    #
    # dfcon = pd.concat([df11, df22], axis='rows')
    # dfg = dfcon.groupby(level=0, axis=0).apply(lambda frame: frame.apply(diff_for_2_vals, threshold=threshold, axis=0))

    dfcopy = df1.copy()
    dfcopy['seq_custom_intenal'] = [x for x in range(0, len(dfcopy))]

    dfg = apply_numerical_threshold(df1, threshold)
    dfd = dfg.iloc[:, 1:].all(1)
    ls_indx = dfd[dfd == True].index.to_list()
    ls_indx = [x - 1 for x in ls_indx] + ls_indx
    return dfcopy[~dfcopy['seq_custom_intenal'].isin(ls_indx)].drop(columns=['seq_custom_intenal'])


def _getCommonDiffs(df1, df2, ls_ref_src, threshold):
    """find the set of rows having differences between two datasets"""

    aa = df2[df2.index.isin(df1.index)].reset_index()
    aa = aa.rename(index={x:y for x,y in zip(aa.index,['Target' for x in aa.index])})
    bb=  df1[df1.index.isin(df2.index)].reset_index()
    bb = bb.rename(index={x:y for x,y in zip(bb.index,['Source' for x in bb.index])})
    
    sortCols = ls_ref_src + ['index']
    cc = pd.concat([aa, bb], sort=False).reset_index().sort_values(sortCols)
    df_cd_full = cc.drop_duplicates(subset=bb.columns, keep=False)
    aa,bb,cc = [None]*3

    df_cd_full = compare_apply_numerical_threshold(df_cd_full, threshold=threshold)
    return df_cd_full

def _reportUniquenessOfReferenceCols(df1,df2):
    """Check if the reference columns have all the values as unique and if not report it"""

    def duplicateDF(ddf,index_cols):
        ddf = ddf.reset_index()[index_cols].sort_values(index_cols)
        return ddf[ddf.duplicated(index_cols, keep = False)].groupby(index_cols).size().reset_index().rename(columns={0:'Count'})
    
    n = 3 #gv.NumOfDuplicateRecordDisplay
    
    isUnique1 = df1.index.is_unique
    if not isUnique1:
        dup = duplicateDF(df1,df1.index.names)
        if len(dup) >n:
            loggerFail(f"Showing only {n} duplicate records in reference column(s) out of {len(dup)} records in File1:")
            # print(f"Showing only {n} duplicate records in reference column(s) out of {len(dup)} records in File1:")
        else:
            loggerFail("Showing all the duplicate records in reference column(s) in dataset1:")
            # print("Showing all the duplicate records in reference column(s) in dataset1:")

        dup_df = dup.head(n).to_string(index=False)
        loggerDisplay(f'{dup_df}')
        loggerDisplay('')
        # print(f'{dup_df}')
        # print('')
        
    isUnique2 = df2.index.is_unique
    if not isUnique2:
        dup = duplicateDF(df2,df2.index.names)
        if len(dup) >n:
            loggerFail(f"Showing only {n} duplicate records in reference column(s) out of {len(dup)} records in File2:")
            # (f"Showing only {n} duplicate records in reference column(s) out of {len(dup)} records in File2:")
        else:
            loggerFail("Showing all the duplicate records in reference column(s) in dataset2:")
            # print("Showing all the duplicate records in reference column(s) in dataset2:")

        dup_df =dup.head(n).to_string(index=False)
        loggerDisplay(f'{dup_df}')
        loggerDisplay('')
        # print(f'{dup_df}')
        # print('')
    
    if isUnique1 == False or isUnique2 == False:
        raise Exception(f'Fix the duplicate values in reference column(s) of source or target dataframe and try again. See details in logger steps.')
        # raise Exception(f'Fix the duplicate values in reference column(s) of source or target dataframe and try again. See details in logger steps. Examples of duplicate reference values\n {dup_df}')
    
    
def _write_df_to_sheet(df, writer, sheet_name, row, col,header=True, index=False):
    """Writes the dataset to given sheet at specified location"""
    try:
        df.to_excel(writer, sheet_name=sheet_name, startrow=row-1, startcol=col-1, header=header, index=index)
    except openpyxl.utils.exceptions.IllegalCharacterError:
        df_changed = df.applymap(lambda x: x.encode('unicode_escape').decode('utf-8') if isinstance(x, str) else x)
        df_changed.to_excel(writer, sheet_name=sheet_name, startrow=row - 1, startcol=col - 1, header=header, index=index)


def _write_in_detailed_report(t_sh, rowtowrite, vtype, df, writer, nameofNewSheet, row, col, msg ):
    """Write in detailed sheet of report"""
    t_sh['A' + str(rowtowrite)] =vtype

    reportdiffSize = get_from_reporting_dict('data_size_in_report') if check_key_in_reporting_dict(
        'data_size_in_report') else default_reportdiffSize

    if len(df)<=reportdiffSize:
        _write_df_to_sheet(df, writer, nameofNewSheet, row =row, col=col, index=False)
    else: # this part is not being executed as df is sliced to reportdiffsize earlier only
        df = df[:reportdiffSize]
        _write_df_to_sheet(df, writer, nameofNewSheet, row =row, col=col, index=False)
        t_sh['B' + str(rowtowrite + reportdiffSize + 1)] =msg
    return df

def apply_numerical_threshold(df, threshold):
    df1 = df.copy()
    df1['seq_custom_intenal'] = [x for x in range(0, len(df1))]  # adding
    df1 = df1.reset_index(drop=True)

    cols_to_drop = [x for x in ['seq_custom_intenal', 'index'] if x in list(df1.columns)]
    df11 = df1[df1.index % 2 != 0].drop(columns=cols_to_drop)

    df22 = df1.shift()
    df22 = df22[df22.index % 2 != 0].drop(columns=cols_to_drop)
    df1 = None  # reset df

    dfcon = pd.concat([df11, df22], axis='rows')
    df11, df22 = [None] * 2  # reset df

    dfg = dfcon.groupby(level=0, axis=0).apply(lambda frame: frame.apply(diff_for_2_vals, threshold=threshold, axis=0))
    if len(dfg)==0:
        dfg = dfcon.drop(index=dfcon.index)

    dfcon = None  # reset df
    return dfg

def getCoordinates_by_apply_numerical_threshold(df, threshold):
    # generate coordinates based on numerical threshold

    dfg = apply_numerical_threshold(df, threshold)
    df_cd1 = dfg.rename(columns={x: y for x, y in zip(dfg.columns, range(1, len(dfg.columns) + 1))})
    diff_cells = df_cd1[df_cd1 == False].stack().index.tolist()
    diff_cells = [(x[0], x[1] + 1) for x in diff_cells]
    return diff_cells


def _getDiffCellsCoordinates(df_cd):
    """Find the cells coordinates of differnces"""
    numeric_threshold = get_from_reporting_dict('numeric_threshold') if check_key_in_reporting_dict('numeric_threshold') else 0

    if numeric_threshold > 0:
        return getCoordinates_by_apply_numerical_threshold(df_cd, numeric_threshold)
    else:
        dd = (df_cd.shift() == df_cd).reset_index()
        ddd= dd[dd.index % 2 != 0]
        df_cd1 = ddd.rename(columns={x:y for x,y in zip(ddd.columns,range(1,len(ddd.columns)+1))})
        dd,ddd = [None]*2
        diff_cells = df_cd1[df_cd1 == False].stack().index.tolist()
        return diff_cells


def _formatExcelCellColor(sh, diff_cells):
    """Colour red for the difference cells in final report"""

    #Open the difference excel and colour red the cells of difference
    DiffColor = "FFFF0000"
    for x,y in diff_cells:
        sh.cell(x+1,y+1).fill= openpyxl.styles.PatternFill(start_color=DiffColor, end_color=DiffColor, fill_type = "solid")
        sh.cell(x+2,y+1).fill= openpyxl.styles.PatternFill(start_color=DiffColor, end_color=DiffColor, fill_type = "solid")
    



    







